import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font as OpenPyXLFont, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import logging
from PyQt6.QtWidgets import QMessageBox
import platform
import re

logger = logging.getLogger(__name__)

class PivotUtils:
    oxide_factors = {
        'Si': ('SiO2', 60.0843 / 28.0855),  # ≈2.1393
        'Ti': ('TiO2', 79.8658 / 47.867),   # ≈1.668
        'Al': ('Al2O3', 101.9613 / (2 * 26.9815)),  # ≈1.8892
        'Fe': ('Fe2O3', 159.6872 / (2 * 55.845)),  # ≈1.4297 (as Fe2O3, common in geochem)
        'Mn': ('MnO', 70.9374 / 54.9380),   # ≈1.2915
        'Mg': ('MgO', 40.3044 / 24.305),    # ≈1.6584
        'Ca': ('CaO', 56.0774 / 40.078),    # ≈1.399
        'Na': ('Na2O', 61.9789 / (2 * 22.9898)),  # ≈1.3473
        'K': ('K2O', 94.1960 / (2 * 39.0983)),  # ≈1.2046
        'P': ('P2O5', 141.9445 / (2 * 30.9738)),  # ≈2.2915
        'Cr': ('Cr2O3', 151.9902 / (2 * 51.9961)),  # ≈1.4604
        'Ni': ('NiO', 74.6928 / 58.6934),   # ≈1.2732
        'Cu': ('CuO', 79.5454 / 63.546),    # ≈1.2523
        'Zn': ('ZnO', 81.3794 / 65.38),     # ≈1.2452
        'Ba': ('BaO', 153.3294 / 137.327),  # ≈1.117
        'Sr': ('SrO', 103.6194 / 87.62),    # ≈1.183
        'V': ('V2O5', 181.8802 / (2 * 50.9415)),  # ≈1.785
        'Zr': ('ZrO2', 123.2182 / 91.224),  # ≈1.351
        'Nb': ('Nb2O5', 265.8098 / (2 * 92.9064)),  # ≈1.430
        'Y': ('Y2O3', 225.8102 / (2 * 88.9058)),  # ≈1.271
        'La': ('La2O3', 325.8092 / (2 * 138.9055)),  # ≈1.172
        'Ce': ('CeO2', 172.1148 / 140.116),  # ≈1.228
        'Nd': ('Nd2O3', 336.482 / (2 * 144.242)),  # ≈1.167
        'Pb': ('PbO', 223.1992 / 207.2),    # ≈1.077
        'Th': ('ThO2', 248.0722 / 232.038),  # ≈1.069
        'U': ('U3O8', 842.088 / (3 * 238.029)),  # ≈1.179 (approximate for U3O8)
    }

    @staticmethod
    def export_pivot(pivot_data, decimal_places):
        """Export pivot table to Excel with styling"""
        if pivot_data is None or pivot_data.empty:
            QMessageBox.warning(None, "Warning", "No data to export!")
            return

        try:
            filtered = pivot_data.copy()

            # Apply filters if needed (simplified for this example)
            # ... (add filter logic if needed)

            file_path = QFileDialog.getSaveFileName(None, "Save Pivot Table", "", "Excel files (*.xlsx)")[0]
            if not file_path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Pivot Table"

            header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            first_col_fill = PatternFill(start_color="FFF5E4", end_color="FFF5E4", fill_type="solid")
            odd_fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
            even_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            header_font = OpenPyXLFont(name="Segoe UI", size=12, bold=True)
            cell_font = OpenPyXLFont(name="Segoe UI", size=12)
            cell_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))

            headers = list(filtered.columns)
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = cell_align
                c.border = thin_border
                ws.column_dimensions[get_column_letter(ci)].width = 15

            dec = int(decimal_places)
            for ri, (_, row) in enumerate(filtered.iterrows(), 2):
                fill = even_fill if (ri - 1) % 2 == 0 else odd_fill
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci)
                    try:
                        cell.value = round(float(val), dec)
                        cell.number_format = f"0.{''.join(['0']*dec)}" if dec > 0 else "0"
                    except Exception:
                        cell.value = "" if pd.isna(val) else str(val)
                    cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = thin_border
                    cell.fill = first_col_fill if ci == 1 else fill

            wb.save(file_path)
            QMessageBox.information(None, "Success", "Pivot table exported successfully!")

            if QMessageBox.question(None, "Open File", "Open the saved Excel file?") == QMessageBox.StandardButton.Yes:
                sys = platform.system()
                try:
                    if sys == "Windows":
                        os.startfile(file_path)
                    elif sys == "Darwin":
                        os.system(f"open '{file_path}'")
                    else:
                        os.system(f"xdg-open '{file_path}'")
                except Exception as e:
                    QMessageBox.warning(None, "Error", f"Failed to open file: {str(e)}")

        except Exception as e:
            QMessageBox.warning(None, "Error", f"Failed to export: {str(e)}")

    @staticmethod
    def check_rm(pivot_tab):
        """Check RM using database (simplified for PyQt6)"""
        # Implement database connection and CRM check
        # This is a placeholder; adapt to your DB setup
        logger.info("CRM check initiated")
        # ... (DB logic here)
        pass

    @staticmethod
    def correct_pivot_crm(pivot_tab, selected_element, max_corr):
        """Correct Pivot CRM values (simplified)"""
        # Implement correction logic
        logger.info(f"Correcting Pivot CRM for {selected_element}")
        # ... (correction logic here)
        pass