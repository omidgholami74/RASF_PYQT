import os
import platform
import pandas as pd
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font as OpenPyXLFont, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class PivotExporter:
    """Handles exporting the pivot table to an Excel file."""
    def __init__(self, pivot_tab):
        self.pivot_tab = pivot_tab
        self.logger = pivot_tab.logger
    
    def export_pivot(self):
        """Export the pivot table to an Excel file with formatting matching the UI."""
        if self.pivot_tab.current_view_df is None or self.pivot_tab.current_view_df.empty:
            self.logger.warning("No data to export")
            QMessageBox.warning(self.pivot_tab, "Warning", "No data to export!")
            return

        try:
            # Verify required methods exist
            if not hasattr(self.pivot_tab, 'is_numeric') or not hasattr(self.pivot_tab, 'format_value'):
                self.logger.error("PivotTab is missing required methods: 'is_numeric' or 'format_value'")
                raise AttributeError("PivotTab is missing required methods: 'is_numeric' or 'format_value'")

            # Get save file path
            file_path = QFileDialog.getSaveFileName(self.pivot_tab, "Save Pivot Table", "pivot_table.xlsx", "Excel files (*.xlsx)")[0]
            if not file_path:
                self.logger.debug("Export cancelled by user")
                self.pivot_tab.status_label.setText("Export cancelled")
                return

            # Prepare data for export
            df = self.pivot_tab.current_view_df.copy()
            try:
                decimal_places = int(self.pivot_tab.decimal_places.currentText())
            except (ValueError, AttributeError) as e:
                self.logger.warning(f"Invalid decimal_places value, using default 1: {str(e)}")
                decimal_places = 1
            
            # Create a list to hold all rows, including CRM and Diff rows
            export_rows = []
            export_index = []

            # Iterate through the displayed rows and insert CRM/Diff rows
            for idx, row in df.iterrows():
                sol_label = row['Solution Label']
                export_rows.append(row)
                export_index.append(sol_label)
                
                # Add inline CRM and Diff rows if they exist
                if sol_label in self.pivot_tab._inline_crm_rows_display:
                    for crm_row, tags in self.pivot_tab._inline_crm_rows_display[sol_label]:
                        if not isinstance(crm_row, list) or not crm_row:
                            self.logger.warning(f"Invalid CRM row data for {sol_label}")
                            continue
                        # Convert CRM/Diff row to a Series with the same columns as df
                        crm_data = pd.Series(index=df.columns, dtype=object)
                        crm_data['Solution Label'] = crm_row[0]
                        for col in df.columns[1:]:  # Skip 'Solution Label'
                            try:
                                col_idx = self.pivot_tab.pivot_data.columns.get_loc(col)
                                crm_data[col] = crm_row[col_idx] if col_idx < len(crm_row) else ''
                            except (ValueError, IndexError) as e:
                                self.logger.warning(f"Invalid column index for {col} in CRM row for {sol_label}: {str(e)}")
                                crm_data[col] = ''
                        if crm_row[0].endswith("CRM"):
                            export_rows.append(crm_data)
                            export_index.append(f"{sol_label} CRM")
                        elif crm_row[0].endswith("Diff (%)") and getattr(self.pivot_tab, 'show_diff', False).isChecked():
                            export_rows.append(crm_data)
                            export_index.append(f"{sol_label} Diff (%)")

            # Create DataFrame for export
            export_df = pd.DataFrame(export_rows, index=export_index)

            # Format numeric columns to respect decimal places
            for col in export_df.columns[1:]:  # Skip 'Solution Label'
                export_df[col] = export_df[col].apply(
                    lambda x: self.pivot_tab.format_value(x) if self.pivot_tab.is_numeric(x) else x
                )

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Pivot Table"

            # Define styles
            header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            first_col_fill = PatternFill(start_color="FFF5E4", end_color="FFF5E4", fill_type="solid")
            odd_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            even_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            diff_in_range_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            diff_out_range_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            header_font = OpenPyXLFont(name="Segoe UI", size=12, bold=True)
            cell_font = OpenPyXLFont(name="Segoe UI", size=12)
            cell_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

            # Write headers
            headers = list(export_df.columns)
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = cell_align
                c.border = thin_border
                ws.column_dimensions[get_column_letter(ci)].width = 15

            # Get difference range
            try:
                min_diff = float(self.pivot_tab.diff_min.text())
                max_diff = float(self.pivot_tab.diff_max.text())
            except (ValueError, AttributeError) as e:
                self.logger.warning(f"Invalid diff_min or diff_max values, using defaults -12 and 12: {str(e)}")
                min_diff, max_diff = -12, 12

            # Write data
            row_idx = 2
            for idx, row in enumerate(export_rows):
                sol_label = export_index[idx]
                is_crm_row = sol_label.endswith("CRM")
                is_diff_row = sol_label.endswith("Diff (%)")
                
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=ci)
                    cell.value = val if is_crm_row or is_diff_row else self.pivot_tab.format_value(val)
                    cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = thin_border
                    # Apply row background
                    cell.fill = first_col_fill if ci == 1 else (even_fill if (row_idx - 1) % 2 == 0 else odd_fill)
                    
                    # Apply difference coloring for non-CRM, non-Diff rows
                    if ci > 1 and not is_crm_row and not is_diff_row and self.pivot_tab.is_numeric(val):
                        col_name = headers[ci - 1]
                        pivot_val = float(val)
                        crm_val = None
                        if sol_label in self.pivot_tab._inline_crm_rows_display:
                            for crm_row, tags in self.pivot_tab._inline_crm_rows_display[sol_label]:
                                if isinstance(crm_row, list) and crm_row and crm_row[0].endswith("CRM"):
                                    try:
                                        crm_val_str = crm_row[self.pivot_tab.pivot_data.columns.get_loc(col_name)] if col_name in self.pivot_tab.pivot_data.columns else ""
                                        if self.pivot_tab.is_numeric(crm_val_str):
                                            crm_val = float(crm_val_str)
                                            break
                                    except ValueError as e:
                                        self.logger.warning(f"Invalid column {col_name} for CRM row in {sol_label}: {str(e)}")
                        if crm_val is not None:
                            diff_percent = ((pivot_val - crm_val) / crm_val * 100) if crm_val != 0 else 0
                            if not (min_diff <= diff_percent <= max_diff):
                                cell.fill = diff_out_range_fill
                            else:
                                cell.fill = diff_in_range_fill
                    # Apply CRM row fill
                    elif is_crm_row and ci > 1:
                        cell.fill = first_col_fill
                    # Apply Diff row fill based on tags
                    elif is_diff_row and ci > 1:
                        col_name = headers[ci - 1]
                        for crm_row, tags in self.pivot_tab._inline_crm_rows_display[sol_label.replace(" Diff (%)", "")]:
                            if isinstance(crm_row, list) and crm_row and crm_row[0].endswith("Diff (%)"):
                                try:
                                    col_idx = self.pivot_tab.pivot_data.columns.get_loc(col_name)
                                    if tags and col_idx - 1 < len(tags):  # Adjust for 0-based tags index
                                        cell.fill = diff_in_range_fill if tags[col_idx - 1] == "in_range" else diff_out_range_fill
                                    else:
                                        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                                except ValueError as e:
                                    self.logger.warning(f"Invalid column {col_name} for Diff row in {sol_label}: {str(e)}")
                                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                                break

                row_idx += 1

            # Adjust column widths
            for ci, col in enumerate(headers, 1):
                max_length = max(
                    len(str(col)),
                    max((len(str(row.get(col, ''))) for row in export_rows), default=10)
                )
                adjusted_width = max_length * 1.2
                ws.column_dimensions[get_column_letter(ci)].width = adjusted_width

            # Save the workbook
            wb.save(file_path)
            self.logger.info(f"Pivot table exported to {file_path}")
            self.pivot_tab.status_label.setText(f"Exported to {file_path}")
            QMessageBox.information(self.pivot_tab, "Success", "Pivot table exported successfully!")
            
            # Ask to open the file
            if QMessageBox.question(self.pivot_tab, "Open File", "Open the saved Excel file?") == QMessageBox.StandardButton.Yes:
                try:
                    if platform.system() == "Windows":
                        os.startfile(file_path)
                    elif platform.system() == "Darwin":
                        os.system(f"open '{file_path}'")
                    else:
                        os.system(f"xdg-open '{file_path}'")
                except Exception as e:
                    self.logger.error(f"Failed to open file: {str(e)}")
                    QMessageBox.warning(self.pivot_tab, "Error", f"Failed to open file: {str(e)}")

        except Exception as e:
            self.logger.error(f"Failed to export pivot table: {str(e)}")
            self.pivot_tab.status_label.setText(f"Error: {str(e)}")
            QMessageBox.warning(self.pivot_tab, "Error", f"Failed to export pivot table: {str(e)}")