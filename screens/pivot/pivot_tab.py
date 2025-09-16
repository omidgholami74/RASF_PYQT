import os
import platform
import re
import pandas as pd
import numpy as np
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QComboBox, QLabel, QTableView,
                             QDialog, QCheckBox, QLineEdit, QFrame, QTreeView, QFileDialog, QMessageBox, QHeaderView)
from PyQt6.QtCore import Qt, QAbstractTableModel, QModelIndex, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QColor, QStandardItemModel, QStandardItem, QPainter, QPen
from PyQt6.QtCharts import QChart, QChartView, QLineSeries, QScatterSeries, QValueAxis, QCategoryAxis
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font as OpenPyXLFont, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sqlite3
import logging

logger = logging.getLogger(__name__)

class CRMWorker(QThread):
    """Worker thread for processing CRM data with optimized database access."""
    finished = pyqtSignal(dict, dict, dict)
    error = pyqtSignal(str)

    def __init__(self, pivot_tab, pivot_data, element_order, oxide_factors, use_int_var, decimal_places):
        super().__init__()
        self.pivot_tab = pivot_tab
        self.pivot_data = pivot_data
        self.element_order = element_order
        self.oxide_factors = oxide_factors
        self.use_int_var = use_int_var
        self.decimal_places = decimal_places
        self._crm_cache = pivot_tab._crm_cache  # استفاده از کش موجود در PivotTab

    def run(self):
        try:
            logger.debug("Starting CRMWorker thread")
            conn = sqlite3.connect(self.pivot_tab.app.crm_tab.db_path)
            cursor = conn.cursor()

            # ایجاد ایندکس برای بهبود عملکرد پرس‌وجو
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_crm_id ON crm ([CRM ID])")

            # بررسی ستون‌های مورد نیاز
            cursor.execute("PRAGMA table_info(crm)")
            cols = [x[1] for x in cursor.fetchall()]
            required = {'CRM ID', 'Element', 'Sort Grade', 'Analysis Method'}
            if not required.issubset(cols):
                self.error.emit("CRM table missing required columns!")
                conn.close()
                return

            # فیلتر کردن ردیف‌های CRM یا par
            crm_rows = self.pivot_data[self.pivot_data['Solution Label'].str.contains('CRM|par', case=False, na=False)]
            if crm_rows.empty:
                self.error.emit("No CRM or par rows found in pivot data!")
                conn.close()
                return

            # استخراج CRM IDها
            crm_ids = set(f"OREAS {m.group()}" for _, row in crm_rows.iterrows() if (m := re.search(r'\d+', str(row['Solution Label']))))

            # اگر داده‌ها در کش موجود باشند، از کش استفاده کن
            if not all(crm_id in self._crm_cache for crm_id in crm_ids):
                placeholders = ','.join(['?'] * len(crm_ids))
                cursor.execute(
                    f"SELECT [CRM ID], [Element], [Sort Grade], [Analysis Method] FROM crm WHERE [CRM ID] IN ({placeholders})",
                    list(crm_ids)
                )
                crm_data = cursor.fetchall()

                for crm_id, element_full, sort_grade, analysis_method in crm_data:
                    symbol = element_full.split(',')[-1].strip() if ',' in element_full else element_full.split()[-1].strip()
                    try:
                        self._crm_cache.setdefault(crm_id, {}).setdefault(analysis_method, {})[symbol] = float(sort_grade)
                    except (ValueError, TypeError):
                        logger.warning(f"Invalid sort_grade for CRM ID {crm_id}, element {symbol}")

            dec = int(self.decimal_places)
            inline_crm_rows = {}
            included_crms = {}
            for _, row in crm_rows.iterrows():
                label = row['Solution Label']
                m = re.search(r'\d+', str(label))
                if not m:
                    continue
                crm_id = f"OREAS {m.group()}"
                analysis_method = '4-Acid Digestion' if 'CRM' in str(label).upper() else 'Aqua Regia Digestion'

                crm_dict = self._crm_cache.get(crm_id, {}).get(analysis_method, {})
                if not crm_dict:
                    continue

                crm_values = {'Solution Label': crm_id}
                for col in self.element_order:
                    if col == 'Solution Label' or col not in self.pivot_data.columns:
                        continue
                    element_symbol = col.split(' ')[0].split('_')[0]
                    if element_symbol in crm_dict:
                        if element_symbol in self.oxide_factors:
                            _, factor = self.oxide_factors[element_symbol]
                            crm_values[col] = crm_dict[element_symbol] * factor
                        else:
                            crm_values[col] = crm_dict[element_symbol]
                if len(crm_values) > 1:
                    inline_crm_rows.setdefault(label, []).append(crm_values)
                    included_crms[label] = QCheckBox(label, checked=True)

            if not inline_crm_rows:
                self.error.emit("No matching CRM elements found for comparison!")
                conn.close()
                return

            inline_crm_rows_display = self.pivot_tab._build_crm_row_lists_for_columns(
                list(self.pivot_data.columns), inline_crm_rows, self._crm_cache
            )
            self.finished.emit(inline_crm_rows, inline_crm_rows_display, included_crms)
            logger.debug("CRMWorker thread finished")
            conn.close()

        except Exception as e:
            logger.error(f"CRMWorker error: {str(e)}")
            self.error.emit(f"Failed to check RM: {str(e)}")
            conn.close()

class FreezeTableWidget(QTableView):
    """A QTableView with a frozen first column, optimized for scrolling."""
    def __init__(self, model, parent=None):
        super().__init__(parent)
        self.frozenTableView = QTableView(self)
        self.setModel(model)
        self.init()

        self.horizontalHeader().sectionResized.connect(self.updateSectionWidth)
        self.verticalHeader().sectionResized.connect(self.updateSectionHeight)
        self.frozenTableView.verticalScrollBar().valueChanged.connect(self.frozenVerticalScroll)
        self.verticalScrollBar().valueChanged.connect(self.mainVerticalScroll)

    def init(self):
        """Initialize the frozen table view with optimized settings."""
        self.frozenTableView.setModel(self.model())
        self.frozenTableView.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.frozenTableView.verticalHeader().hide()
        self.frozenTableView.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.viewport().stackUnder(self.frozenTableView)

        stylesheet = """
            QTableView {
                border: 1px solid #cccccc;
                selection-background-color: #999;
                gridline-color: #cccccc;
                background-color: white;
            }
            QTableView::item {
                border: 1px solid #cccccc;
                padding: 4px;
            }
        """
        self.setStyleSheet(stylesheet)
        self.frozenTableView.setStyleSheet(stylesheet)
        self.frozenTableView.setSelectionModel(self.selectionModel())

        for col in range(1, self.model().columnCount()):
            self.frozenTableView.setColumnHidden(col, True)

        default_width = self.columnWidth(0) if self.columnWidth(0) > 0 else 150
        self.frozenTableView.setColumnWidth(0, default_width)
        self.frozenTableView.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.frozenTableView.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.frozenTableView.show()

        self.setHorizontalScrollMode(QTableView.ScrollMode.ScrollPerPixel)
        self.setVerticalScrollMode(QTableView.ScrollMode.ScrollPerPixel)
        self.frozenTableView.setVerticalScrollMode(QTableView.ScrollMode.ScrollPerPixel)

        # غیرفعال کردن به‌روزرسانی‌های غیرضروری
        self.setUpdatesEnabled(False)
        self.frozenTableView.setUpdatesEnabled(False)
        self.updateFrozenTableGeometry()
        self.setUpdatesEnabled(True)
        self.frozenTableView.setUpdatesEnabled(True)

    def updateSectionWidth(self, logicalIndex, oldSize, newSize):
        if logicalIndex == 0:
            self.frozenTableView.setColumnWidth(0, newSize)
            self.updateFrozenTableGeometry()

    def updateSectionHeight(self, logicalIndex, oldSize, newSize):
        self.frozenTableView.setRowHeight(logicalIndex, newSize)

    def frozenVerticalScroll(self, value):
        self.setUpdatesEnabled(False)
        self.verticalScrollBar().setValue(value)
        self.setUpdatesEnabled(True)

    def mainVerticalScroll(self, value):
        self.frozenTableView.setUpdatesEnabled(False)
        self.frozenTableView.verticalScrollBar().setValue(value)
        self.frozenTableView.setUpdatesEnabled(True)

    def updateFrozenTableGeometry(self):
        self.frozenTableView.setGeometry(
            self.verticalHeader().width() + self.frameWidth(),
            self.frameWidth(),
            self.columnWidth(0) or 150,
            self.viewport().height() + self.horizontalHeader().height()
        )

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.updateFrozenTableGeometry()

    def moveCursor(self, cursorAction, modifiers):
        current = super().moveCursor(cursorAction, modifiers)
        if cursorAction == QTableView.CursorAction.MoveLeft and current.column() > 0:
            visual_x = self.visualRect(current).topLeft().x()
            if visual_x < self.frozenTableView.columnWidth(0):
                new_value = self.horizontalScrollBar().value() + visual_x - self.frozenTableView.columnWidth(0)
                self.horizontalScrollBar().setValue(int(new_value))
        return current

    def scrollTo(self, index, hint=QTableView.ScrollHint.EnsureVisible):
        if index.column() > 0:
            super().scrollTo(index, hint)

class PivotTableModel(QAbstractTableModel):
    """Custom table model optimized for large datasets."""
    def __init__(self, pivot_tab, df=None, crm_rows=None):
        super().__init__()
        self.pivot_tab = pivot_tab
        self._df = pd.DataFrame() if df is None else df
        self._crm_rows = {} if crm_rows is None else crm_rows
        self._row_mapping = []
        self._column_widths = {}
        self._cache = {}  # کش برای داده‌های رندر شده
        self.set_data(df, crm_rows)

    def set_data(self, df, crm_rows=None):
        # logger.debug(f"Setting data with {len(df)} rows, CRM rows: {len(crm_rows)}")
        self.beginResetModel()
        self._df = df.copy() if df is not None else pd.DataFrame()
        self._crm_rows = crm_rows if crm_rows is not None else {}
        self._row_mapping = self._build_row_mapping()
        self._cache.clear()  # پاک کردن کش هنگام تغییر داده‌ها
        self.endResetModel()

    def _build_row_mapping(self):
        mapping = []
        for idx, sol_label in enumerate(self._df['Solution Label']):
            mapping.append(('pivot', idx))
            if sol_label in self._crm_rows:
                for i, _ in enumerate(self._crm_rows[sol_label]):
                    mapping.append(('crm' if i == 0 else 'diff', sol_label, i))
        logger.debug(f"Built row mapping: {len(mapping)} rows")
        return mapping

    def rowCount(self, parent=QModelIndex()):
        return len(self._row_mapping)

    def columnCount(self, parent=QModelIndex()):
        return self._df.shape[1] if not self._df.empty else 0

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid() or not (0 <= index.row() < self.rowCount() and 0 <= index.column() < self.columnCount()):
            return None

        row = index.row()
        col = index.column()
        col_name = self._df.columns[col]
        row_type, data_idx, *extra = self._row_mapping[row]

        # استفاده از کش برای داده‌های نمایشی
        cache_key = (row, col, role)
        if cache_key in self._cache:
            return self._cache[cache_key]

        if role == Qt.ItemDataRole.DisplayRole:
            dec = int(self.pivot_tab.decimal_places.currentText())
            if row_type == 'pivot':
                value = self._df.iloc[data_idx, col]
                if pd.isna(value):
                    result = ""
                elif col_name == "Solution Label":
                    result = str(value)
                else:
                    try:
                        result = f"{float(value):.{dec}f}"
                    except (ValueError, TypeError):
                        result = str(value)
            elif row_type == 'crm':
                sol_label, crm_idx = extra
                if sol_label not in self._crm_rows or crm_idx >= len(self._crm_rows[sol_label]):
                    result = ""
                else:
                    row_data, _ = self._crm_rows[sol_label][crm_idx]
                    value = row_data[col]
                    result = str(value) if value else ""
            elif row_type == 'diff':
                sol_label, crm_idx = extra
                if sol_label not in self._crm_rows or crm_idx >= len(self._crm_rows[sol_label]):
                    result = ""
                else:
                    row_data, _ = self._crm_rows[sol_label][crm_idx]
                    value = row_data[col]
                    result = str(value) if value else ""
            self._cache[cache_key] = result
            return result

        elif role == Qt.ItemDataRole.BackgroundRole:
            if row_type == 'crm':
                result = QColor("#FFF5E4")
            elif row_type == 'diff':
                sol_label, crm_idx = extra
                if sol_label not in self._crm_rows or crm_idx >= len(self._crm_rows[sol_label]):
                    result = QColor("#E6E6FA")
                else:
                    _, tags = self._crm_rows[sol_label][crm_idx]
                    result = QColor("#ECFFC4") if tags[col] == "in_range" else \
                             QColor("#FFCCCC") if tags[col] == "out_range" else \
                             QColor("#E6E6FA")
            else:
                result = QColor("#f9f9f9") if data_idx % 2 == 0 else QColor("white")
            self._cache[cache_key] = result
            return result

        elif role == Qt.ItemDataRole.TextAlignmentRole:
            result = Qt.AlignmentFlag.AlignLeft if col_name == "Solution Label" else Qt.AlignmentFlag.AlignCenter
            self._cache[cache_key] = result
            return result

        return None

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal and section < self.columnCount():
                return str(self._df.columns[section])
            return str(section + 1)
        return None

    def set_column_width(self, col, width):
        self._column_widths[col] = width

class FilterDialog(QDialog):
    """Dialog for row/column filtering."""
    def __init__(self, parent, title, is_row_filter=True):
        super().__init__(parent)
        self.parent = parent
        self.is_row_filter = is_row_filter
        self.setWindowTitle(title)
        self.setGeometry(200, 200, 300, 460)
        self.setModal(True)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Select Field:"))

        self.field_combo = QComboBox()
        if self.is_row_filter:
            self.field_combo.addItems(["Solution Label", "Element"])
        else:
            self.field_combo.addItems(["Element"])
        layout.addWidget(self.field_combo)

        self.tree_view = QTreeView()
        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels(["Value", "Select"])
        self.tree_view.setModel(self.model)
        self.tree_view.setRootIsDecorated(False)
        self.tree_view.header().setStretchLastSection(False)
        self.tree_view.header().resizeSection(0, 160)
        self.tree_view.header().resizeSection(1, 80)
        layout.addWidget(self.tree_view)

        button_layout = QHBoxLayout()
        select_all_btn = QPushButton("Select All")
        select_all_btn.clicked.connect(lambda: self.set_all_checks(True))
        button_layout.addWidget(select_all_btn)

        deselect_all_btn = QPushButton("Deselect All")
        deselect_all_btn.clicked.connect(lambda: self.set_all_checks(False))
        button_layout.addWidget(deselect_all_btn)

        apply_btn = QPushButton("Apply")
        apply_btn.clicked.connect(self.accept)
        button_layout.addWidget(apply_btn)

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.reject)
        button_layout.addWidget(close_btn)

        layout.addLayout(button_layout)
        self.field_combo.currentTextChanged.connect(self.update_tree)
        self.update_tree()

    def update_tree(self):
        self.model.clear()
        self.model.setHorizontalHeaderLabels(["Value", "Select"])
        field = self.field_combo.currentText()
        if not field:
            return

        var_dict = self.parent.row_filter_values if self.is_row_filter else self.parent.column_filter_values
        uniques = self.parent.solution_label_order if self.is_row_filter and field == "Solution Label" else \
                  sorted(self.parent.pivot_data[field].astype(str).unique()) if field in self.parent.pivot_data.columns else \
                  self.parent.element_order
        if field not in var_dict:
            var_dict[field] = {v: QCheckBox(v, checked=True) for v in uniques}

        for val in uniques:
            value_item = QStandardItem(val)
            check_item = QStandardItem()
            check_item.setCheckable(True)
            check_item.setCheckState(Qt.CheckState.Checked if var_dict[field][val].isChecked() else Qt.CheckState.Unchecked)
            self.model.appendRow([value_item, check_item])

        self.tree_view.clicked.connect(self.toggle_check)

    def toggle_check(self, index):
        if index.column() != 1:
            return
        val = self.model.item(index.row(), 0).text()
        field = self.field_combo.currentText()
        var_dict = self.parent.row_filter_values if self.is_row_filter else self.parent.column_filter_values
        if field in var_dict and val in var_dict[field]:
            var_dict[field][val].setChecked(not var_dict[field][val].isChecked())
            self.model.item(index.row(), 1).setCheckState(
                Qt.CheckState.Checked if var_dict[field][val].isChecked() else Qt.CheckState.Unchecked)
            self.parent.update_pivot_display()

    def set_all_checks(self, value):
        field = self.field_combo.currentText()
        var_dict = self.parent.row_filter_values if self.is_row_filter else self.parent.column_filter_values
        if field not in var_dict:
            return
        for val, checkbox in var_dict[field].items():
            checkbox.setChecked(value)
        self.update_tree()
        self.parent.update_pivot_display()

class PivotPlotDialog(QDialog):
    """Dialog for plotting CRM data with a separate difference plot."""
    def __init__(self, parent, selected_element):
        super().__init__(parent)
        self.parent = parent
        self.selected_element = selected_element
        self.setWindowTitle(f"CRM Plot for {selected_element}")
        self.setGeometry(200, 200, 1000, 800)
        self.setModal(False)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        control_frame = QHBoxLayout()
        self.show_check_crm = QCheckBox("Show Check CRM", checked=True)
        self.show_check_crm.toggled.connect(self.update_plot)
        control_frame.addWidget(self.show_check_crm)

        self.show_pivot_crm = QCheckBox("Show Pivot CRM", checked=True)
        self.show_pivot_crm.toggled.connect(self.update_plot)
        control_frame.addWidget(self.show_pivot_crm)

        self.show_middle = QCheckBox("Show Middle", checked=True)
        self.show_middle.toggled.connect(self.update_plot)
        control_frame.addWidget(self.show_middle)

        self.show_range = QCheckBox("Show Range", checked=True)
        self.show_range.toggled.connect(self.update_plot)
        control_frame.addWidget(self.show_range)

        self.show_diff = QCheckBox("Show Difference", checked=True)
        self.show_diff.toggled.connect(self.update_plot)
        control_frame.addWidget(self.show_diff)

        control_frame.addWidget(QLabel("Max Correction (%):"))
        self.max_correction_percent = QLineEdit("32")
        control_frame.addWidget(self.max_correction_percent)

        correct_btn = QPushButton("Correct Pivot CRM")
        correct_btn.clicked.connect(self.parent.correct_pivot_crm)
        control_frame.addWidget(correct_btn)

        select_crms_btn = QPushButton("Select CRMs")
        select_crms_btn.clicked.connect(self.open_select_crms_window)
        control_frame.addWidget(select_crms_btn)

        layout.addLayout(control_frame)

        self.chart = QChart()
        self.chart_view = QChartView(self.chart)
        self.chart_view.setRenderHint(QPainter.RenderHint.Antialiasing)
        layout.addWidget(self.chart_view)

        self.diff_chart = QChart()
        self.diff_chart_view = QChartView(self.diff_chart)
        self.diff_chart_view.setRenderHint(QPainter.RenderHint.Antialiasing)
        layout.addWidget(self.diff_chart_view)

        self.update_plot()

    def update_plot(self):
        if not self.selected_element or self.selected_element not in self.parent.pivot_data.columns:
            QMessageBox.warning(self, "Warning", f"Element '{self.selected_element}' not found in pivot data!")
            return

        try:
            self.chart.removeAllSeries()
            self.diff_chart.removeAllSeries()
            for axis in self.chart.axes() + self.diff_chart.axes():
                self.chart.removeAxis(axis)
                self.diff_chart.removeAxis(axis)

            valid_pairs = []
            for sol_label, crm_rows in self.parent._inline_crm_rows_display.items():
                if sol_label not in self.parent.included_crms or not self.parent.included_crms[sol_label].isChecked():
                    continue
                pivot_row = self.parent.pivot_data[self.parent.pivot_data['Solution Label'] == sol_label]
                if pivot_row.empty:
                    continue
                pivot_val = pivot_row.iloc[0][self.selected_element]
                for row_data, _ in crm_rows:
                    if isinstance(row_data, list) and row_data and row_data[0].endswith("CRM"):
                        val = row_data[self.parent.pivot_data.columns.get_loc(self.selected_element)] if self.selected_element in self.parent.pivot_data.columns else ""
                        if val and val.strip():
                            try:
                                crm_val = float(val)
                                if pd.notna(pivot_val):
                                    pivot_val_float = float(pivot_val)
                                    valid_pairs.append((sol_label, crm_val, pivot_val_float))
                            except ValueError:
                                continue

            valid_pairs.sort(key=lambda x: x[0])
            crm_display_values = [pair[1] for pair in valid_pairs]
            crm_pivot_values = [pair[2] for pair in valid_pairs]
            labels = [pair[0] for pair in valid_pairs]

            if not valid_pairs:
                pivot_data = self.parent.pivot_data[self.parent.pivot_data['Solution Label'].str.contains('CRM|par', case=False, na=False)]
                if not pivot_data.empty and self.show_pivot_crm.isChecked():
                    series = QLineSeries()
                    series.setName(f"{self.selected_element} (Pivot CRM)")
                    series.setPen(QPen(QColor("blue"), 2))
                    for i, (_, row) in enumerate(pivot_data.iterrows()):
                        val = row[self.selected_element]
                        try:
                            if pd.notna(val):
                                series.append(i, float(val))
                        except ValueError:
                            continue
                    if series.count() > 0:
                        self.chart.addSeries(series)
                        axis_x = QCategoryAxis()
                        for i, (_, row) in enumerate(pivot_data.iterrows()):
                            axis_x.append(row['Solution Label'], i)
                        axis_x.setLabelsAngle(45)
                        axis_y = QValueAxis()
                        axis_y.setTitleText(f"{self.selected_element} Value")
                        self.chart.addAxis(axis_x, Qt.AlignmentFlag.AlignBottom)
                        self.chart.addAxis(axis_y, Qt.AlignmentFlag.AlignLeft)
                        series.attachAxis(axis_x)
                        series.attachAxis(axis_y)
                        self.chart.setTitle(f"Pivot CRM Values for {self.selected_element}")
                        self.chart_view.update()
                        self.diff_chart.setTitle("No difference data available")
                        self.diff_chart_view.update()
                        return
                self.chart.setTitle(f"No valid CRM data for {self.selected_element}")
                self.chart_view.update()
                self.diff_chart.setTitle("No difference data available")
                self.diff_chart_view.update()
                QMessageBox.warning(self, "Warning", f"No valid CRM data for {self.selected_element}")
                return

            middle_values = [(disp_val + piv_val) / 2 for disp_val, piv_val in zip(crm_display_values, crm_pivot_values)]
            range_lower = [val - self.parent.calculate_dynamic_range(val) for val in crm_display_values]
            range_upper = [val + self.parent.calculate_dynamic_range(val) for val in crm_display_values]
            diff_values = [((crm_val - piv_val) / crm_val) * 100 if crm_val != 0 else 0
                           for crm_val, piv_val in zip(crm_display_values, crm_pivot_values)]

            axis_x = QCategoryAxis()
            for i, label in enumerate(labels):
                axis_x.append(label, i)
            axis_x.setLabelsAngle(45)
            self.chart.addAxis(axis_x, Qt.AlignmentFlag.AlignBottom)

            axis_y = QValueAxis()
            axis_y.setTitleText(f"{self.selected_element} Value")
            self.chart.addAxis(axis_y, Qt.AlignmentFlag.AlignLeft)

            axis_x_diff = QCategoryAxis()
            for i, label in enumerate(labels):
                axis_x_diff.append(label, i)
            axis_x_diff.setLabelsAngle(45)
            self.diff_chart.addAxis(axis_x_diff, Qt.AlignmentFlag.AlignBottom)

            axis_y_diff = QValueAxis()
            axis_y_diff.setTitleText("Difference (%)")
            axis_y_diff.setLabelFormat("%.1f")
            self.diff_chart.addAxis(axis_y_diff, Qt.AlignmentFlag.AlignLeft)

            plotted = False
            if self.show_check_crm.isChecked() and any(crm_display_values):
                series = QLineSeries()
                series.setName(f"{self.selected_element} (Check CRM)")
                series.setPen(QPen(QColor("red"), 2, Qt.PenStyle.DashLine))
                for i, val in enumerate(crm_display_values):
                    series.append(i, val)
                self.chart.addSeries(series)
                series.attachAxis(axis_x)
                series.attachAxis(axis_y)
                if self.show_range.isChecked():
                    range_series = QLineSeries()
                    range_series.setName("Check CRM Range")
                    range_series.setPen(QPen(QColor("red"), 1, Qt.PenStyle.DotLine))
                    for i, (lower, upper) in enumerate(zip(range_lower, range_upper)):
                        range_series.append(i, lower)
                        range_series.append(i, upper)
                    self.chart.addSeries(range_series)
                    range_series.attachAxis(axis_x)
                    range_series.attachAxis(axis_y)
                plotted = True

            if self.show_pivot_crm.isChecked() and any(crm_pivot_values):
                series = QLineSeries()
                series.setName(f"{self.selected_element} (Pivot CRM)")
                series.setPen(QPen(QColor("blue"), 2))
                scatter = QScatterSeries()
                scatter.setMarkerShape(QScatterSeries.MarkerShape.MarkerShapeCircle)
                scatter.setMarkerSize(8)
                scatter.setPen(QPen(QColor("blue")))
                for i, val in enumerate(crm_pivot_values):
                    series.append(i, val)
                    scatter.append(i, val)
                self.chart.addSeries(series)
                self.chart.addSeries(scatter)
                series.attachAxis(axis_x)
                series.attachAxis(axis_y)
                scatter.attachAxis(axis_x)
                series.attachAxis(axis_y)
                plotted = True

            if self.show_middle.isChecked() and any(middle_values):
                series = QLineSeries()
                series.setName(f"{self.selected_element} (Middle)")
                series.setPen(QPen(QColor("green"), 2, Qt.PenStyle.DotLine))
                for i, val in enumerate(middle_values):
                    series.append(i, val)
                self.chart.addSeries(series)
                series.attachAxis(axis_x)
                series.attachAxis(axis_y)
                plotted = True

            if self.show_diff.isChecked() and any(diff_values):
                series = QLineSeries()
                series.setName("Difference (%)")
                series.setPen(QPen(QColor("purple"), 2))
                scatter = QScatterSeries()
                scatter.setMarkerShape(QScatterSeries.MarkerShape.MarkerShapeCircle)
                scatter.setMarkerSize(8)
                scatter.setPen(QPen(QColor("purple")))
                for i, val in enumerate(diff_values):
                    series.append(i, val)
                    scatter.append(i, val)
                self.diff_chart.addSeries(series)
                self.diff_chart.addSeries(scatter)
                series.attachAxis(axis_x_diff)
                series.attachAxis(axis_y_diff)
                scatter.attachAxis(axis_x_diff)
                series.attachAxis(axis_y_diff)
                plotted = True

            if not plotted:
                self.chart.setTitle(f"No data plotted for {self.selected_element}")
                self.diff_chart.setTitle("No difference data plotted")
                self.chart_view.update()
                self.diff_chart_view.update()
                return

            self.chart.setTitle(f"CRM Values for {self.selected_element}")
            self.diff_chart.setTitle(f"Difference (%) for {self.selected_element}")
            self.chart_view.update()
            self.diff_chart_view.update()

        except Exception as e:
            logger.error(f"Plot update error: {str(e)}")
            QMessageBox.warning(self, "Error", f"Failed to update plot: {str(e)}")

    def open_select_crms_window(self):
        w = QDialog(self)
        w.setWindowTitle("Select CRMs to Include")
        w.setGeometry(200, 200, 300, 400)
        w.setModal(True)
        layout = QVBoxLayout(w)

        tree_view = QTreeView()
        model = QStandardItemModel()
        model.setHorizontalHeaderLabels(["Label", "Include"])
        tree_view.setModel(model)
        tree_view.setRootIsDecorated(False)
        tree_view.header().resizeSection(0, 160)
        tree_view.header().resizeSection(1, 80)

        for label in sorted(self.parent.included_crms.keys()):
            value_item = QStandardItem(label)
            check_item = QStandardItem()
            check_item.setCheckable(True)
            check_item.setCheckState(Qt.CheckState.Checked if self.parent.included_crms[label].isChecked() else Qt.CheckState.Unchecked)
            model.appendRow([value_item, check_item])

        tree_view.clicked.connect(lambda index: self.toggle_crm_check(index, model))
        layout.addWidget(tree_view)

        button_layout = QHBoxLayout()
        select_all_btn = QPushButton("Select All")
        select_all_btn.clicked.connect(lambda: self.set_all_crms(True, model))
        button_layout.addWidget(select_all_btn)

        deselect_all_btn = QPushButton("Deselect All")
        deselect_all_btn.clicked.connect(lambda: self.set_all_crms(False, model))
        button_layout.addWidget(deselect_all_btn)

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(w.accept)
        button_layout.addWidget(close_btn)

        layout.addLayout(button_layout)
        w.exec()

    def toggle_crm_check(self, index, model):
        if index.column() != 1:
            return
        label = model.item(index.row(), 0).text()
        if label in self.parent.included_crms:
            self.parent.included_crms[label].setChecked(not self.parent.included_crms[label].isChecked())
            model.item(index.row(), 1).setCheckState(
                Qt.CheckState.Checked if self.parent.included_crms[label].isChecked() else Qt.CheckState.Unchecked)
            self.update_plot()

    def set_all_crms(self, value, model):
        for label, checkbox in self.parent.included_crms.items():
            checkbox.setChecked(value)
        model.clear()
        model.setHorizontalHeaderLabels(["Label", "Include"])
        for label in sorted(self.parent.included_crms.keys()):
            value_item = QStandardItem(label)
            check_item = QStandardItem()
            check_item.setCheckable(True)
            check_item.setCheckState(Qt.CheckState.Checked if self.parent.included_crms[label].isChecked() else Qt.CheckState.Unchecked)
            model.appendRow([value_item, check_item])
        self.update_plot()

class PivotTab(QWidget):
    """PivotTab with inline CRM rows, difference coloring, and plot visualization."""
    oxide_factors = {
        'Si': ('SiO2', 60.0843 / 28.0855),
        'Ti': ('TiO2', 79.8658 / 47.867),
        'Al': ('Al2O3', 101.9613 / (2 * 26.9815)),
        'Fe': ('Fe2O3', 159.6872 / (2 * 55.845)),
        'Mn': ('MnO', 70.9374 / 54.9380),
        'Mg': ('MgO', 40.3044 / 24.305),
        'Ca': ('CaO', 56.0774 / 40.078),
        'Na': ('Na2O', 61.9789 / (2 * 22.9898)),
        'K': ('K2O', 94.1960 / (2 * 39.0983)),
        'P': ('P2O5', 141.9445 / (2 * 30.9738)),
        'Cr': ('Cr2O3', 151.9902 / (2 * 51.9961)),
        'Ni': ('NiO', 74.6928 / 58.6934),
        'Cu': ('CuO', 79.5454 / 63.546),
        'Zn': ('ZnO', 81.3794 / 65.38),
        'Ba': ('BaO', 153.3294 / 137.327),
        'Sr': ('SrO', 103.6194 / 87.62),
        'V': ('V2O5', 181.8802 / (2 * 50.9415)),
        'Zr': ('ZrO2', 123.2182 / 91.224),
        'Nb': ('Nb2O5', 265.8098 / (2 * 92.9064)),
        'Y': ('Y2O3', 225.8102 / (2 * 88.9058)),
        'La': ('La2O3', 325.8092 / (2 * 138.9055)),
        'Ce': ('CeO2', 172.1148 / 140.116),
        'Nd': ('Nd2O3', 336.482 / (2 * 144.242)),
        'Pb': ('PbO', 223.1992 / 207.2),
        'Th': ('ThO2', 248.0722 / 232.038),
        'U': ('U3O8', 842.088 / (3 * 238.029)),
    }

    def __init__(self, app, parent_frame):
        super().__init__(parent_frame)
        self.app = app
        self.parent_frame = parent_frame
        self.pivot_data = None
        self.solution_label_order = None
        self.element_order = None
        self.row_filter_values = {}
        self.column_filter_values = {}
        self.original_df = None
        self.column_widths = {}
        self.cached_formatted = {}
        self.current_view_df = None
        self._inline_crm_rows = {}
        self._inline_crm_rows_display = {}
        self._crm_cache = {}  # کش برای داده‌های CRM
        self.search_var = QLineEdit()
        self.row_filter_field = QComboBox()
        self.column_filter_field = QComboBox()
        self.element_selector = QComboBox()
        self.decimal_places = QComboBox()
        self.use_int_var = QCheckBox("Use Int")
        self.diff_min = QLineEdit("-12")
        self.diff_max = QLineEdit("12")
        self.show_check_crm = QCheckBox("Show Check CRM", checked=True)
        self.show_pivot_crm = QCheckBox("Show Pivot CRM", checked=True)
        self.show_middle = QCheckBox("Show Middle", checked=True)
        self.show_diff = QCheckBox("Show Diff", checked=True)
        self.show_range = QCheckBox("Show Range", checked=True)
        self.max_correction_percent = QLineEdit("32")
        self.included_crms = {}
        self.worker = None
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        control_frame = QFrame()
        control_layout = QHBoxLayout(control_frame)

        control_layout.addWidget(QLabel("Decimal Places:"))
        self.decimal_places.addItems(["0", "1", "2", "3"])
        self.decimal_places.setCurrentText("1")
        self.decimal_places.currentTextChanged.connect(self.update_pivot_display)
        control_layout.addWidget(self.decimal_places)

        self.use_int_var.toggled.connect(self.create_pivot)
        control_layout.addWidget(self.use_int_var)

        control_layout.addWidget(QLabel("Diff Range (%):"))
        self.diff_min.textChanged.connect(self.validate_diff_range)
        control_layout.addWidget(self.diff_min)
        control_layout.addWidget(QLabel("to"))
        self.diff_max.textChanged.connect(self.validate_diff_range)
        control_layout.addWidget(self.diff_max)

        control_layout.addWidget(QLabel("Select Element:"))
        self.element_selector.currentTextChanged.connect(self.show_element_plot)
        control_layout.addWidget(self.element_selector)

        self.plot_btn = QPushButton("Show Plot")
        self.plot_btn.clicked.connect(self.show_element_plot)
        control_layout.addWidget(self.plot_btn)

        self.search_var.setPlaceholderText("Search...")
        self.search_var.textChanged.connect(self.update_pivot_display)
        control_layout.addWidget(self.search_var)

        row_filter_btn = QPushButton("Row Filter")
        row_filter_btn.clicked.connect(self.open_row_filter_window)
        control_layout.addWidget(row_filter_btn)

        col_filter_btn = QPushButton("Column Filter")
        col_filter_btn.clicked.connect(self.open_column_filter_window)
        control_layout.addWidget(col_filter_btn)

        self.check_rm_btn = QPushButton("Check RM")
        self.check_rm_btn.clicked.connect(self.check_rm)
        control_layout.addWidget(self.check_rm_btn)

        clear_crm_btn = QPushButton("Clear CRM")
        clear_crm_btn.clicked.connect(self.clear_inline_crm)
        control_layout.addWidget(clear_crm_btn)

        export_btn = QPushButton("Export")
        export_btn.clicked.connect(self.export_pivot)
        control_layout.addWidget(export_btn)

        layout.addWidget(control_frame)

        self.table_view = FreezeTableWidget(PivotTableModel(self))
        self.table_view.setAlternatingRowColors(True)
        self.table_view.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.table_view.setSortingEnabled(True)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table_view.doubleClicked.connect(self.on_cell_double_click)
        layout.addWidget(self.table_view)

        self.status_label = QLabel("Pivot table will be displayed here.")
        self.status_label.setFont(QFont("Segoe UI", 14))
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.status_label)

    def validate_diff_range(self):
        try:
            min_val = float(self.diff_min.text())
            max_val = float(self.diff_max.text())
            if min_val > max_val:
                self.diff_min.setText(str(max_val - 1))
            self.update_pivot_display()
        except ValueError:
            pass

    def create_pivot(self):
        df = self.app.get_data()
        if df is None or df.empty:
            QMessageBox.warning(self, "Warning", "No data to display!")
            return

        try:
            logger.debug(f"Creating pivot with data shape: {df.shape}")
            self.original_df = df.copy()
            df_filtered = df[df['Type'].isin(['Samp', 'Sample'])].copy()
            df_filtered['original_index'] = df_filtered.index
            df_filtered['Element'] = df_filtered['Element'].str.split('_').str[0]
            df_filtered['unique_id'] = df_filtered.groupby(['Solution Label', 'Element']).cumcount()

            self.solution_label_order = df_filtered['Solution Label'].drop_duplicates().tolist()
            self.element_order = df_filtered['Element'].drop_duplicates().tolist()
            self.element_selector.clear()
            self.element_selector.addItems([""] + self.element_order)

            value_column = 'Int' if self.use_int_var.isChecked() else 'Corr Con'
            if value_column not in df_filtered.columns:
                QMessageBox.warning(self, "Error", f"Column '{value_column}' not found in data!")
                return

            pivot_df = df_filtered.pivot_table(
                index=['Solution Label', 'unique_id'],
                columns='Element',
                values=value_column,
                aggfunc='first'
            ).reset_index()
            pivot_df = pivot_df.merge(
                df_filtered[['original_index', 'Solution Label', 'unique_id']],
                on=['Solution Label', 'unique_id'],
                how='left'
            ).sort_values('original_index').drop(columns=['original_index', 'unique_id']).drop_duplicates()

            self.pivot_data = pivot_df
            logger.debug(f"Pivot data created with shape: {pivot_df.shape}")
            self.column_widths.clear()
            self.cached_formatted.clear()
            self._inline_crm_rows.clear()
            self._inline_crm_rows_display.clear()
            self._crm_cache.clear()
            self.update_pivot_display()

        except Exception as e:
            logger.error(f"Pivot creation error: {str(e)}")
            QMessageBox.warning(self, "Pivot Error", f"Failed to create pivot table: {str(e)}")

    def update_pivot_display(self):
        if self.pivot_data is None or self.pivot_data.empty:
            self.status_label.setText("No data loaded")
            self.table_view.setModel(None)
            return

        self.table_view.setUpdatesEnabled(False)
        df = self.pivot_data.copy()
        s = self.search_var.text().strip().lower()
        if s:
            mask = df.apply(lambda r: r.astype(str).str.lower().str.contains(s, na=False).any(), axis=1)
            df = df[mask]

        rf = self.row_filter_field.currentText()
        if rf and rf in self.row_filter_values:
            selected = [k for k, v in self.row_filter_values[rf].items() if v.isChecked()]
            if selected:
                df = df[df[rf].isin(selected)]

        cf = self.column_filter_field.currentText()
        if cf and cf in self.column_filter_values:
            selected_cols = [k for k, v in self.column_filter_values[cf].items() if v.isChecked()]
            if selected_cols:
                keep = ['Solution Label'] + [c for c in self.element_order if c in selected_cols and c in df.columns]
                df = df[keep]

        self.current_view_df = df.reset_index(drop=True)
        crm_rows = {sol_label: data for sol_label, data in self._inline_crm_rows_display.items() if sol_label in df['Solution Label'].values}
        model = PivotTableModel(self, df, crm_rows)
        self.table_view.setModel(model)

        for col in range(model.columnCount()):
            self.table_view.setColumnWidth(col, 100)
            self.table_view.frozenTableView.setColumnWidth(col, 100)
        self.table_view.resizeRowsToContents()
        self.table_view.frozenTableView.resizeRowsToContents()
        self.table_view.frozenTableView.update()
        self.table_view.setUpdatesEnabled(True)

        self.status_label.setText(f"Data loaded: {len(df)} rows")
        logger.debug(f"Updated pivot display with {len(df)} rows, {len(crm_rows)} CRM rows")

    def _build_crm_row_lists_for_columns(self, columns, inline_crm_rows, crm_cache):
        crm_display = {}
        dec = int(self.decimal_places.currentText())
        try:
            min_diff = float(self.diff_min.text())
            max_diff = float(self.diff_max.text())
        except ValueError:
            min_diff, max_diff = -12, 12

        element_params = {}
        if self.original_df is not None:
            sample_rows = self.original_df[self.original_df['Type'].isin(['Sample', 'Samp'])]
            for element in self.element_order:
                rows = sample_rows[sample_rows['Element'].str.startswith(element)]
                if not rows.empty:
                    row = rows.iloc[0]
                    element_params[element] = {
                        'Act Vol': row.get('Act Vol', 1.0),
                        'Act Wgt': row.get('Act Wgt', 1.0),
                        'Coeff 1': row.get('Coeff 1', 0.0),
                        'Coeff 2': row.get('Coeff 2', 1.0)
                    }

        for sol_label, list_of_dicts in inline_crm_rows.items():
            if sol_label not in self.pivot_data['Solution Label'].values:
                continue
            crm_display[sol_label] = []
            pivot_row = self.pivot_data[self.pivot_data['Solution Label'] == sol_label]
            if pivot_row.empty:
                continue
            pivot_values = pivot_row.iloc[0].to_dict()

            for d in list_of_dicts:
                crm_row_list = [''] * len(columns)
                crm_row_list[0] = f"{d.get('Solution Label', sol_label)} CRM"
                diff_row_list = [''] * len(columns)
                diff_row_list[0] = f"{sol_label} Diff (%)"
                diff_tags = ['diff'] * len(columns)

                for col_idx, col in enumerate(columns[1:], 1):
                    val = d.get(col, "")
                    element_symbol = col.split('_')[0]
                    params = element_params.get(element_symbol, {
                        'Act Vol': 1.0, 'Act Wgt': 1.0, 'Coeff 1': 0.0, 'Coeff 2': 1.0
                    })
                    try:
                        if pd.notna(val) and val != "":
                            if self.use_int_var.isChecked() and params['Act Vol'] != 0 and params['Act Wgt'] != 0:
                                int_value = params['Coeff 2'] * (float(val) / (params['Act Vol'] / params['Act Wgt'])) + params['Coeff 1']
                                if element_symbol in self.oxide_factors:
                                    _, factor = self.oxide_factors[element_symbol]
                                    crm_row_list[col_idx] = f"{int_value * factor:.{dec}f}"
                                else:
                                    crm_row_list[col_idx] = f"{int_value:.{dec}f}"
                            else:
                                if element_symbol in self.oxide_factors:
                                    _, factor = self.oxide_factors[element_symbol]
                                    crm_row_list[col_idx] = f"{float(val) * factor:.{dec}f}"
                                else:
                                    crm_row_list[col_idx] = f"{float(val):.{dec}f}"

                        pivot_val = pivot_values.get(col, None)
                        if pivot_val is not None and val is not None and pd.notna(pivot_val) and pd.notna(val):
                            pivot_val = float(pivot_val)
                            if element_symbol in self.oxide_factors:
                                _, pivot_factor = self.oxide_factors[element_symbol]
                                pivot_val *= pivot_factor
                            if self.use_int_var.isChecked() and params['Act Vol'] != 0 and params['Act Wgt'] != 0:
                                crm_val = params['Coeff 2'] * (float(val) / (params['Act Vol'] / params['Act Wgt'])) + params['Coeff 1']
                            else:
                                crm_val = float(val)
                            if crm_val != 0:
                                diff = ((crm_val - pivot_val) / crm_val) * 100
                                diff_row_list[col_idx] = f"{diff:.{dec}f}"
                                diff_tags[col_idx] = "in_range" if min_diff <= diff <= max_diff else "out_range"
                    except (ValueError, TypeError) as e:
                        logger.debug(f"Error processing CRM value for {col}: {str(e)}")

                crm_display[sol_label].append((crm_row_list, ['crm'] * len(columns)))
                crm_display[sol_label].append((diff_row_list, diff_tags))

        logger.debug(f"Built CRM display for {len(crm_display)} solution labels")
        return crm_display

    def calculate_dynamic_range(self, value):
        try:
            value = float(value)
            return value * (0.2 if value < 100 else 0.1 if value <= 1000 else 0.05)
        except (ValueError, TypeError):
            return 0

    def on_cell_double_click(self, index):
        if not index.isValid() or self.current_view_df is None:
            return
        row = index.row()
        col = index.column()
        col_name = self.current_view_df.columns[col]
        if col_name == "Solution Label":
            return

        try:
            row_type, data_row, *extra = self.table_view.model()._row_mapping[row]
            if row_type in ('crm', 'diff'):
                return
            solution_label = self.current_view_df.iloc[data_row]['Solution Label']
            element = col_name.split('_')[0]
            cond = (self.original_df['Solution Label'] == solution_label) & (self.original_df['Element'].str.startswith(element))
            cond &= (self.original_df['Type'] == 'Samp')
            match = self.original_df[cond]
            if match.empty:
                return

            r = match.iloc[0]
            value_column = 'Int' if self.use_int_var.isChecked() else 'Corr Con'
            value = float(r.get(value_column, 0)) / 10000
            info = [
                f"Solution: {solution_label}",
                f"Element: {col_name}",
                f"Act Wgt: {self.format_value(r.get('Act Wgt', 'N/A'))}",
                f"Act Vol: {self.format_value(r.get('Act Vol', 'N/A'))}",
                f"DF: {self.format_value(r.get('DF', 'N/A'))}",
                f"Concentration: {self.format_value(value)}"
            ]
            if element.split()[0] in self.oxide_factors:
                formula, factor = self.oxide_factors[element.split()[0]]
                try:
                    oxide_value = float(value) * factor
                    info.extend([f"Oxide Formula: {formula}", f"Oxide %: {self.format_value(oxide_value)}"])
                except (ValueError, TypeError):
                    info.extend([f"Oxide Formula: {formula}", "Oxide %: N/A"])

            w = QDialog(self)
            w.setWindowTitle("Cell Information")
            w.setGeometry(200, 200, 300, 200)
            layout = QVBoxLayout(w)
            for line in info:
                layout.addWidget(QLabel(line))
            close_btn = QPushButton("Close")
            close_btn.clicked.connect(w.accept)
            layout.addWidget(close_btn)
            w.exec()

        except Exception as e:
            logger.error(f"Cell double-click error: {str(e)}")
            QMessageBox.warning(self, "Error", f"Failed to display cell info: {str(e)}")

    def format_value(self, x):
        try:
            d = int(self.decimal_places.currentText())
            return f"{float(x):.{d}f}"
        except (ValueError, TypeError):
            return "" if pd.isna(x) else str(x)

    def open_row_filter_window(self):
        if self.pivot_data is None:
            QMessageBox.warning(self, "Warning", "No data to filter!")
            return
        FilterDialog(self, "Row Filter", is_row_filter=True).exec()

    def open_column_filter_window(self):
        if self.pivot_data is None:
            QMessageBox.warning(self, "Warning", "No data to filter!")
            return
        FilterDialog(self, "Column Filter", is_row_filter=False).exec()

    def export_pivot(self):
        if self.pivot_data is None or self.pivot_data.empty:
            QMessageBox.warning(self, "Warning", "No data to export!")
            return

        try:
            filtered = self.pivot_data.copy()
            rf = self.row_filter_field.currentText()
            if rf and rf in self.row_filter_values:
                selected = [k for k, v in self.row_filter_values[rf].items() if v.isChecked()]
                if selected:
                    filtered = filtered[filtered[rf].isin(selected)]

            cf = self.column_filter_field.currentText()
            if cf and cf in self.column_filter_values:
                selected_cols = [k for k, v in self.column_filter_values[cf].items() if v.isChecked()]
                if selected_cols:
                    keep = ['Solution Label'] + [c for c in self.element_order if c in selected_cols and c in filtered.columns]
                    filtered = filtered[keep]

            filtered['Solution Label'] = pd.Categorical(filtered['Solution Label'], categories=self.solution_label_order, ordered=True)
            filtered = filtered.sort_values('Solution Label').reset_index(drop=True)

            file_path = QFileDialog.getSaveFileName(self, "Save Pivot Table", "", "Excel files (*.xlsx)")[0]
            if not file_path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Pivot Table"

            header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            first_col_fill = PatternFill(start_color="FFF5E4", end_color="FFF5E4", fill_type="solid")
            odd_fill = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
            even_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            diff_in_range_fill = PatternFill(start_color="ECFFC4", end_color="ECFFC4", fill_type="solid")
            diff_out_range_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            header_font = OpenPyXLFont(name="Segoe UI", size=12, bold=True)
            cell_font = OpenPyXLFont(name="Segoe UI", size=12)
            cell_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

            headers = list(filtered.columns)
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = cell_align
                c.border = thin_border
                ws.column_dimensions[get_column_letter(ci)].width = 15

            dec = int(self.decimal_places.currentText())
            try:
                min_diff = float(self.diff_min.text())
                max_diff = float(self.diff_max.text())
            except ValueError:
                min_diff, max_diff = -12, 12

            row_idx = 2
            for _, row in filtered.iterrows():
                sol_label = row['Solution Label']
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=ci)
                    try:
                        cell.value = round(float(val), dec)
                        cell.number_format = f"0.{''.join(['0']*dec)}" if dec > 0 else "0"
                    except (ValueError, TypeError):
                        cell.value = "" if pd.isna(val) else str(val)
                    cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = thin_border
                    cell.fill = first_col_fill if ci == 1 else (even_fill if (row_idx - 1) % 2 == 0 else odd_fill)
                row_idx += 1

                if sol_label in self._inline_crm_rows_display:
                    for crm_row_data, tags in self._inline_crm_rows_display[sol_label]:
                        for ci, val in enumerate(crm_row_data, 1):
                            cell = ws.cell(row=row_idx, column=ci)
                            cell.value = str(val) if val else ""
                            cell.font = cell_font
                            cell.alignment = cell_align
                            cell.border = thin_border
                            if crm_row_data[0].endswith("CRM"):
                                cell.fill = first_col_fill if ci == 1 else PatternFill(start_color="FFF5E4", end_color="FFF5E4", fill_type="solid")
                            elif crm_row_data[0].endswith("Diff (%)"):
                                if tags[ci-1] == "in_range":
                                    cell.fill = diff_in_range_fill
                                elif tags[ci-1] == "out_range":
                                    cell.fill = diff_out_range_fill
                                else:
                                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                        row_idx += 1

            wb.save(file_path)
            QMessageBox.information(self, "Success", "Pivot table exported successfully!")
            if QMessageBox.question(self, "Open File", "Open the saved Excel file?") == QMessageBox.StandardButton.Yes:
                try:
                    if platform.system() == "Windows":
                        os.startfile(file_path)
                    elif platform.system() == "Darwin":
                        os.system(f"open '{file_path}'")
                    else:
                        os.system(f"xdg-open '{file_path}'")
                except Exception as e:
                    logger.error(f"Export file open error: {str(e)}")
                    QMessageBox.warning(self, "Error", f"Failed to open file: {str(e)}")

        except Exception as e:
            logger.error(f"Export error: {str(e)}")
            QMessageBox.warning(self, "Error", f"Failed to export: {str(e)}")

    def reset_cache(self):
        self.pivot_data = None
        self.solution_label_order = None
        self.element_order = None
        self.column_widths.clear()
        self.cached_formatted.clear()
        self.original_df = None
        self._inline_crm_rows.clear()
        self._inline_crm_rows_display.clear()
        self._crm_cache.clear()
        if self.worker is not None:
            self.worker.quit()
            self.worker.wait()
            self.worker = None

    def clear_inline_crm(self):
        self._inline_crm_rows.clear()
        self._inline_crm_rows_display.clear()
        self.included_crms.clear()
        self._crm_cache.clear()
        self.update_pivot_display()

    def check_rm(self):
        if self.pivot_data is None or self.pivot_data.empty:
            QMessageBox.warning(self, "Warning", "No pivot data available!")
            return

        if self.worker is not None and self.worker.isRunning():
            QMessageBox.warning(self, "Warning", "CRM processing is already in progress!")
            return

        self.status_label.setText("Processing CRM data...")
        self.check_rm_btn.setEnabled(False)
        self.worker = CRMWorker(self, self.pivot_data, self.element_order, self.oxide_factors, self.use_int_var.isChecked(), self.decimal_places.currentText())
        self.worker.finished.connect(self.on_crm_processed)
        self.worker.error.connect(self.on_crm_error)
        self.worker.finished.connect(self.worker.deleteLater)
        self.worker.start()

    def on_crm_processed(self, inline_crm_rows, inline_crm_rows_display, included_crms):
        self._inline_crm_rows = inline_crm_rows
        self._inline_crm_rows_display = inline_crm_rows_display
        self.included_crms = included_crms
        self.update_pivot_display()
        self.status_label.setText("CRM data loaded successfully")
        self.check_rm_btn.setEnabled(True)
        self.worker = None
        logger.debug("CRM data processed and UI updated")

    def on_crm_error(self, error_msg):
        QMessageBox.warning(self, "Error", error_msg)
        self.status_label.setText("Failed to load CRM data")
        self.check_rm_btn.setEnabled(True)
        self.worker = None
        logger.debug("CRM processing failed")

    def correct_pivot_crm(self):
        if self.pivot_data is None or self.pivot_data.empty:
            QMessageBox.warning(self, "Warning", "No pivot data available!")
            return

        selected_element = self.element_selector.currentText()
        if not selected_element:
            QMessageBox.warning(self, "Warning", "Please select an element!")
            return

        try:
            max_corr = float(self.max_correction_percent.text()) / 100
            self.pivot_data[selected_element] = pd.to_numeric(self.pivot_data[selected_element], errors='coerce')
            ratios = []
            problematic_labels = []

            for sol_label, crm_rows in self._inline_crm_rows_display.items():
                if sol_label not in self.included_crms or not self.included_crms[sol_label].isChecked():
                    continue
                pivot_row = self.pivot_data[self.pivot_data['Solution Label'] == sol_label]
                if pivot_row.empty:
                    continue
                pivot_val = pivot_row.iloc[0][selected_element]
                for row_data, _ in crm_rows:
                    if isinstance(row_data, list) and row_data and row_data[0].endswith("CRM"):
                        val = row_data[self.pivot_data.columns.get_loc(selected_element)] if selected_element in self.pivot_data.columns else ""
                        if val and val.strip() and pd.notna(pivot_val):
                            try:
                                check_val = float(val)
                                pivot_val = float(pivot_val)
                                range_val = self.calculate_dynamic_range(check_val)
                                lower, upper = check_val - range_val, check_val + range_val
                                if pivot_val < lower:
                                    ratio = lower / pivot_val if pivot_val != 0 else float('inf')
                                    if abs(ratio - 1) <= max_corr:
                                        ratios.append(ratio)
                                        self.pivot_data.loc[self.pivot_data['Solution Label'] == sol_label, selected_element] = pivot_val * ratio
                                    else:
                                        problematic_labels.append(sol_label)
                                elif pivot_val > upper:
                                    ratio = upper / pivot_val if pivot_val != 0 else float('inf')
                                    if abs(ratio - 1) <= max_corr:
                                        ratios.append(ratio)
                                        self.pivot_data.loc[self.pivot_data['Solution Label'] == sol_label, selected_element] = pivot_val * ratio
                                    else:
                                        problematic_labels.append(sol_label)
                            except ValueError:
                                continue

            if problematic_labels:
                QMessageBox.warning(self, "Warning", f"CRM has problem in these points (correction exceeds max):\n" + "\n".join(problematic_labels))

            if not ratios:
                QMessageBox.warning(self, "Warning", "No valid ratios for correction!")
                return

            avg_ratio = np.mean(ratios)
            self.pivot_data[selected_element] = self.pivot_data[selected_element].apply(
                lambda x: x * avg_ratio if pd.notna(x) else x
            )
            self.update_pivot_display()
            QMessageBox.information(self, "Success", f"Pivot CRM corrected for {selected_element} with average ratio={avg_ratio:.3f}")

        except Exception as e:
            logger.error(f"Correct pivot CRM error: {str(e)}")
            QMessageBox.warning(self, "Error", f"Failed to correct Pivot CRM: {str(e)}")

    def show_element_plot(self):
        if self.pivot_data is None or self.pivot_data.empty:
            QMessageBox.warning(self, "Warning", "No data to plot!")
            return

        selected_element = self.element_selector.currentText()
        if not selected_element or selected_element not in self.pivot_data.columns:
            QMessageBox.warning(self, "Warning", f"Element '{selected_element}' not found in pivot data!")
            return

        crm_labels = [label for label in self._inline_crm_rows_display.keys() if 'CRM' in label or 'par' in label]
        if not self.included_crms:
            self.included_crms = {label: QCheckBox(label, checked=True) for label in crm_labels}

        PivotPlotDialog(self, selected_element).exec()