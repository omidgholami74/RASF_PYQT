import sys
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QTableView,QAbstractItemView,
    QHeaderView, QScrollBar, QComboBox, QLineEdit, QDialog, QFileDialog, QMessageBox, QGroupBox
)
from PyQt6.QtCore import Qt, QAbstractTableModel, QVariant, QTimer
from PyQt6.QtGui import QStandardItemModel, QStandardItem, QFont
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import numpy as np
import os
import platform
import logging

# Setup logging
logger = logging.getLogger(__name__)

# Global stylesheet for consistent UI
global_style = """
    QWidget {
        background-color: #F5F7FA;
        font-family: 'Inter', 'Segoe UI', sans-serif;
        font-size: 13px;
    }
    QDialog {
        background-color: #F5F7FA;
    }
    QGroupBox {
        font-weight: bold;
        color: #1A3C34;
        margin-top: 15px;
        border: 1px solid #D0D7DE;
        border-radius: 6px;
        padding: 10px;
    }
    QGroupBox::title {
        subcontrol-origin: margin;
        subcontrol-position: top left;
        padding: 0 5px;
        left: 10px;
    }
    QPushButton {
        background-color: #2E7D32;
        color: white;
        border: none;
        padding: 8px 16px;
        font-weight: 600;
        font-size: 13px;
        border-radius: 6px;
    }
    QPushButton:hover {
        background-color: #1B5E20;
    }
    QPushButton:disabled {
        background-color: #E0E0E0;
        color: #6B7280;
    }
    QComboBox {
        background-color: #FFFFFF;
        color: #1A3C34;
        border: 1px solid #D0D7DE;
        padding: 6px;
        font-size: 13px;
        border-radius: 6px;
    }
    QComboBox:focus {
        border: 1px solid #2E7D32;
        box-shadow: 0 0 5px rgba(46, 125, 50, 0.3);
    }
    QComboBox QAbstractItemView {
        background-color: #FFFFFF;
        color: #1A3C34;
        selection-background-color: #DBEAFE;
        selection-color: #1A3C34;
    }
    QLineEdit {
        background-color: #FFFFFF;
        color: #1A3C34;
        border: 1px solid #D0D7DE;
        padding: 6px;
        font-size: 13px;
        border-radius: 6px;
    }
    QLineEdit:focus {
        border: 1px solid #2E7D32;
        box-shadow: 0 0 5px rgba(46, 125, 50, 0.3);
    }
    QTableView {
        background-color: #FFFFFF;
        border: 1px solid #D0D7DE;
        gridline-color: #E5E7EB;
        font-size: 12px;
        selection-background-color: #DBEAFE;
        selection-color: #1A3C34;
    }
    QHeaderView::section {
        background-color: #F9FAFB;
        font-weight: 600;
        color: #1A3C34;
        border: 1px solid #D0D7DE;
        padding: 6px;
    }
    QTableView::item {
        border: 1px solid #D0D7DE;
    }
    QTableView::item:selected {
        background-color: #DBEAFE;
        color: #1A3C34;
    }
"""

class PandasModel(QAbstractTableModel):
    """Custom model to display pandas DataFrame in QTableView"""
    def __init__(self, data=pd.DataFrame()):
        super().__init__()
        self._data = data

    def rowCount(self, parent=None):
        return len(self._data)

    def columnCount(self, parent=None):
        return len(self._data.columns)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return QVariant()
        if role == Qt.ItemDataRole.DisplayRole:
            value = self._data.iloc[index.row(), index.column()]
            return str(value)
        elif role == Qt.ItemDataRole.BackgroundRole:
            return QColor("#F9FAFB") if index.row() % 2 else Qt.GlobalColor.white
        return QVariant()

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                return str(self._data.columns[section])
            return str(self._data.index[section])
        return QVariant()

class FreezeTableWidget(QTableView):
    """Custom QTableView with a frozen first column"""
    def __init__(self, model, parent=None):
        super().__init__(parent)
        self.frozenTableView = QTableView(self)
        self.setModel(model)
        self.frozenTableView.setModel(model)
        self.init()

        self.horizontalHeader().sectionResized.connect(self.updateSectionWidth)
        self.verticalHeader().sectionResized.connect(self.updateSectionHeight)
        self.frozenTableView.verticalScrollBar().valueChanged.connect(self.frozenVerticalScroll)
        self.verticalScrollBar().valueChanged.connect(self.mainVerticalScroll)

    def init(self):
        self.frozenTableView.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.frozenTableView.verticalHeader().hide()
        self.frozenTableView.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.viewport().stackUnder(self.frozenTableView)
        self.frozenTableView.setStyleSheet(global_style)
        self.frozenTableView.setSelectionModel(self.selectionModel())
        
        for col in range(self.model().columnCount()):
            self.frozenTableView.setColumnHidden(col, col != 0)
        self.frozenTableView.setColumnWidth(0, self.columnWidth(0))
        self.frozenTableView.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.frozenTableView.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerItem)
        self.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerItem)
        self.frozenTableView.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerItem)
        self.updateFrozenTableGeometry()
        self.frozenTableView.show()
        self.frozenTableView.viewport().repaint()

    def updateSectionWidth(self, logicalIndex, oldSize, newSize):
        if logicalIndex == 0:
            self.frozenTableView.setColumnWidth(0, newSize)
            self.updateFrozenTableGeometry()
            self.frozenTableView.viewport().repaint()

    def updateSectionHeight(self, logicalIndex, oldSize, newSize):
        self.frozenTableView.setRowHeight(logicalIndex, newSize)
        self.frozenTableView.viewport().repaint()

    def frozenVerticalScroll(self, value):
        self.verticalScrollBar().setValue(value)
        self.frozenTableView.viewport().repaint()
        self.viewport().update()

    def mainVerticalScroll(self, value):
        self.frozenTableView.verticalScrollBar().setValue(value)
        self.frozenTableView.viewport().repaint()
        self.viewport().update()

    def updateFrozenTableGeometry(self):
        self.frozenTableView.setGeometry(
            self.verticalHeader().width() + self.frameWidth(),
            self.frameWidth(),
            self.columnWidth(0),
            self.viewport().height() + self.horizontalHeader().height()
        )
        self.frozenTableView.viewport().repaint()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.updateFrozenTableGeometry()
        self.frozenTableView.viewport().repaint()

    def moveCursor(self, cursorAction, modifiers):
        current = super().moveCursor(cursorAction, modifiers)
        if cursorAction == QAbstractItemView.CursorAction.MoveLeft and current.column() > 0:
            visual_x = self.visualRect(current).topLeft().x()
            if visual_x < self.frozenTableView.columnWidth(0):
                new_value = self.horizontalScrollBar().value() + visual_x - self.frozenTableView.columnWidth(0)
                self.horizontalScrollBar().setValue(int(new_value))
        return current

    def scrollTo(self, index, hint=QAbstractItemView.ScrollHint.EnsureVisible):
        if index.column() > 0:
            super().scrollTo(index, hint)
        self.frozenTableView.viewport().repaint()

class FilterDialog(QDialog):
    """Dialog for selecting filter values"""
    def __init__(self, parent, filter_values, filter_field, solution_label_order, element_order, update_callback):
        super().__init__(parent)
        self.setWindowTitle("Filter Pivot Table")
        self.setFixedSize(350, 500)
        self.filter_values = filter_values
        self.filter_field = filter_field
        self.solution_label_order = solution_label_order
        self.element_order = element_order
        self.update_callback = update_callback
        self.setStyleSheet(global_style)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        filter_group = QGroupBox("Select Filter Column")
        filter_layout = QVBoxLayout(filter_group)
        filter_layout.setSpacing(10)
        self.filter_combo = QComboBox()
        self.filter_combo.addItems(["Solution Label", "Element"])
        self.filter_combo.setCurrentText(self.filter_field)
        self.filter_combo.setToolTip("Choose the column to filter by")
        self.filter_combo.currentTextChanged.connect(self.update_checkboxes)
        filter_layout.addWidget(self.filter_combo)
        layout.addWidget(filter_group)

        self.filter_table = QTableView()
        self.filter_table.setSelectionMode(QTableView.SelectionMode.NoSelection)
        self.filter_table.setStyleSheet(global_style)
        self.filter_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.filter_table.setToolTip("Select values to include in the pivot table")
        layout.addWidget(self.filter_table, stretch=1)

        button_group = QGroupBox("Filter Actions")
        button_layout = QHBoxLayout(button_group)
        button_layout.setSpacing(12)

        select_all_btn = QPushButton("Select All")
        select_all_btn.setToolTip("Select all filter values")
        select_all_btn.clicked.connect(lambda: self.set_all_checkboxes(True))
        button_layout.addWidget(select_all_btn)

        deselect_all_btn = QPushButton("Deselect All")
        deselect_all_btn.setToolTip("Deselect all filter values")
        deselect_all_btn.clicked.connect(lambda: self.set_all_checkboxes(False))
        button_layout.addWidget(deselect_all_btn)

        close_btn = QPushButton("Close")
        close_btn.setToolTip("Close the filter window (changes are applied automatically)")
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn)

        layout.addWidget(button_group)
        self.update_checkboxes()

    def update_checkboxes(self):
        start_time = time.time()
        field = self.filter_combo.currentText()
        model = QStandardItemModel()
        model.setHorizontalHeaderLabels(["Value", "Select"])

        unique_values = (
            self.solution_label_order if field == "Solution Label"
            else self.element_order if self.element_order else []
        )
        if field not in self.filter_values:
            self.filter_values[field] = {val: True for val in unique_values}

        for value in unique_values:
            value_item = QStandardItem(str(value))
            value_item.setEditable(False)
            check_item = QStandardItem()
            check_item.setCheckable(True)
            check_item.setCheckState(
                Qt.CheckState.Checked if self.filter_values[field].get(value, True)
                else Qt.CheckState.Unchecked
            )
            model.appendRow([value_item, check_item])

        self.filter_table.setModel(model)
        self.filter_table.setColumnWidth(0, 200)
        self.filter_table.setColumnWidth(1, 100)
        model.itemChanged.connect(lambda item: self.toggle_filter(item, field))
        logger.debug(f"Updated checkboxes for {field} in {time.time() - start_time:.3f} seconds")

    def toggle_filter(self, item, field):
        if item.column() == 1:
            value = self.filter_table.model().item(item.row(), 0).text()
            self.filter_values[field][value] = (item.checkState() == Qt.CheckState.Checked)
            QTimer.singleShot(0, self.update_callback)
            logger.debug(f"Toggled filter for {field}: {value} = {self.filter_values[field][value]}")

    def set_all_checkboxes(self, value):
        start_time = time.time()
        field = self.filter_combo.currentText()
        if field in self.filter_values:
            for val in self.filter_values[field]:
                self.filter_values[field][val] = value
            self.update_checkboxes()
            QTimer.singleShot(0, self.update_callback)
            logger.debug(f"Set all checkboxes for {field} to {value} in {time.time() - start_time:.3f} seconds")

class ResultsFrame(QWidget):
    def __init__(self, app, parent=None):
        super().__init__(parent)
        self.app = app
        self.setStyleSheet(global_style)
        self.search_var = ""
        self.filter_field = "Solution Label"
        self.filter_values = {}
        self.search_window = None
        self.column_widths = {}
        self.last_filtered_data = None
        self._last_cache_key = None
        self.solution_label_order = None
        self.element_order = None
        self.decimal_places = "1"
        self.data_hash = None  # Track data changes
        self.setup_ui()
        self.app.notify_data_changed = self.show_processed_data

    def setup_ui(self):
        """Setup Results UI with pivot table, scrollbars, and decimal places selection"""
        start_time = time.time()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        controls_group = QGroupBox("Table Controls")
        controls_layout = QHBoxLayout(controls_group)
        controls_layout.setSpacing(12)
        controls_layout.setContentsMargins(10, 10, 10, 10)

        search_button = QPushButton("🔍 Search")
        search_button.setToolTip("Search the pivot table")
        search_button.clicked.connect(self.open_search_window)
        search_button.setFixedWidth(120)
        controls_layout.addWidget(search_button)

        filter_button = QPushButton("📌 Filter")
        filter_button.setToolTip("Filter the pivot table by Solution Label or Element")
        filter_button.clicked.connect(self.open_filter_window)
        filter_button.setFixedWidth(120)
        controls_layout.addWidget(filter_button)

        self.save_button = QPushButton("💾 Save Excel")
        self.save_button.setToolTip("Save the pivot table to an Excel file")
        self.save_button.clicked.connect(self.save_processed_excel)
        self.save_button.setFixedWidth(120)
        controls_layout.addWidget(self.save_button)

        decimal_label = QLabel("Decimal Places:")
        decimal_label.setFont(QFont("Inter", 12))
        controls_layout.addWidget(decimal_label)
        self.decimal_combo = QComboBox()
        self.decimal_combo.addItems(["0", "1", "2", "3"])
        self.decimal_combo.setCurrentText(self.decimal_places)
        self.decimal_combo.setFixedWidth(60)
        self.decimal_combo.setToolTip("Set the number of decimal places for numeric values")
        self.decimal_combo.currentTextChanged.connect(lambda: QTimer.singleShot(0, self.show_processed_data))
        controls_layout.addWidget(self.decimal_combo)
        controls_layout.addStretch()

        layout.addWidget(controls_group)

        table_group = QGroupBox("Pivot Table")
        table_layout = QVBoxLayout(table_group)
        table_layout.setContentsMargins(0, 0, 0, 0)

        self.processed_table = FreezeTableWidget(PandasModel())
        self.processed_table.setStyleSheet(global_style)
        self.processed_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.processed_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.processed_table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.processed_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.processed_table.setToolTip("Processed pivot table with filtered data")
        table_layout.addWidget(self.processed_table)

        layout.addWidget(table_group, stretch=1)
        self.setLayout(layout)

        logger.debug(f"ResultsFrame UI setup took {time.time() - start_time:.3f} seconds")

    def format_value(self, x):
        """Format a value to the specified number of decimal places"""
        try:
            decimal_places = int(self.decimal_combo.currentText())
            return f"{float(x):.{decimal_places}f}"
        except (ValueError, TypeError):
            return str(x)

    def get_filtered_data(self):
        """Get filtered pivot table data with optimized caching"""
        start_time = time.time()
        df = self.app.get_data()
        if df is None or df.empty:
            logger.warning("No data available in get_filtered_data")
            return None

        required_columns = ['Solution Label', 'Element', 'Corr Con', 'Type']
        if not all(col in df.columns for col in required_columns):
            logger.error(f"DataFrame missing required columns: {required_columns}")
            QMessageBox.warning(self, "Error", f"DataFrame missing required columns: {required_columns}")
            return None

        # Calculate data hash to detect changes
        new_hash = str(pd.util.hash_pandas_object(df[required_columns]).sum())
        if new_hash == self.data_hash and self.last_filtered_data is not None:
            logger.debug(f"Using cached data (same hash), took {time.time() - start_time:.3f} seconds")
            return self.last_filtered_data

        df_filtered = df[df['Type'].isin(['Samp', 'Sample', 'RM', 'Std'])].copy()
        df_filtered = df_filtered[
            (~df_filtered['Solution Label'].isin(self.app.get_excluded_samples())) &
            (~df_filtered['Solution Label'].isin(self.app.get_excluded_volumes())) &
            (~df_filtered['Solution Label'].isin(self.app.get_excluded_dfs()))
        ]

        if df_filtered.empty:
            logger.warning("No data after filtering in get_filtered_data")
            return None

        if 'original_index' not in df_filtered.columns:
            df_filtered['original_index'] = df_filtered.index
        df_filtered['Element'] = df_filtered['Element'].str.split('_').str[0]

        if 'row_id' in df_filtered.columns:
            df_filtered['unique_id'] = df_filtered['row_id']
        else:
            df_filtered['unique_id'] = df_filtered.groupby(['Solution Label', 'Element']).cumcount()

        if self.solution_label_order is None or not self.solution_label_order:
            self.solution_label_order = df_filtered['Solution Label'].drop_duplicates().tolist()
        if self.element_order is None or not self.element_order:
            self.element_order = df_filtered['Element'].drop_duplicates().tolist()

        value_column = 'Corr Con'
        if value_column not in df_filtered.columns:
            logger.error(f"Column '{value_column}' not found in data")
            QMessageBox.warning(self, "Error", f"Column '{value_column}' not found in data!")
            return None

        pivot_data = df_filtered.pivot_table(
            index=['Solution Label', 'unique_id'],
            columns='Element',
            values=value_column,
            aggfunc='first'
        ).reset_index()

        pivot_data = pivot_data.merge(
            df_filtered[['Solution Label', 'unique_id', 'original_index']].drop_duplicates(),
            on=['Solution Label', 'unique_id'],
            how='left'
        ).sort_values('original_index').drop(columns=['unique_id'])

        columns_to_keep = ['Solution Label'] + [col for col in self.element_order if col in pivot_data.columns]
        pivot_data = pivot_data[columns_to_keep]

        search_text = self.search_var.lower().strip()
        filter_field = self.filter_field
        selected_values = [k for k, v in self.filter_values.get(filter_field, {}).items() if v]

        cache_key = (
            search_text,
            filter_field,
            tuple(sorted(selected_values)),
            new_hash
        )
        if cache_key == self._last_cache_key and self.last_filtered_data is not None:
            logger.debug(f"Using cached data, took {time.time() - start_time:.3f} seconds")
            return self.last_filtered_data

        if search_text:
            mask = pivot_data.astype(str).apply(lambda x: x.str.lower().str.contains(search_text, na=False)).any(axis=1)
            pivot_data = pivot_data[mask]

        if filter_field and selected_values:
            if filter_field == 'Solution Label':
                selected_order = [x for x in self.solution_label_order if x in selected_values and x in pivot_data['Solution Label'].values]
                pivot_data = pivot_data[pivot_data['Solution Label'].isin(selected_values)]
                if not pivot_data.empty:
                    pivot_data['Solution Label'] = pd.Categorical(
                        pivot_data['Solution Label'],
                        categories=selected_order,
                        ordered=True
                    )
                    pivot_data = pivot_data.sort_values('Solution Label').reset_index(drop=True)
            elif filter_field == 'Element':
                columns_to_keep = ['Solution Label'] + [col for col in self.element_order if col in selected_values and col in pivot_data.columns]
                pivot_data = pivot_data[columns_to_keep]

        pivot_data = pivot_data.drop_duplicates().reset_index(drop=True)
        self.last_filtered_data = pivot_data
        self._last_cache_key = cache_key
        self.data_hash = new_hash
        logger.debug(f"Get filtered data took {time.time() - start_time:.3f} seconds")
        return pivot_data

    def show_processed_data(self):
        """Show the final processed pivot table"""
        start_time = time.time()
        df = self.get_filtered_data()
        model = QStandardItemModel()

        if df is None or df.empty:
            model.setHorizontalHeaderLabels(["Status"])
            model.appendRow([QStandardItem("No data loaded")])
            self.processed_table.setModel(model)
            self.processed_table.setColumnWidth(0, 150)
            self.column_widths = {"Status": 150}
            logger.debug(f"Show processed data (no data) took {time.time() - start_time:.3f} seconds")
            return

        columns = list(df.columns)
        model.setHorizontalHeaderLabels(columns)
        self.column_widths = {}

        for col_idx, col in enumerate(columns):
            max_width = max(
                [len(str(col))] + [len(self.format_value(x)) for x in df[col].dropna().head(100)],
                default=10
            )
            pixel_width = min(max_width * 10, 150)
            self.column_widths[col] = pixel_width
            self.processed_table.setColumnWidth(col_idx, pixel_width)

        for row in df.itertuples(index=False):
            items = [QStandardItem(self.format_value(x)) for x in row]
            for item in items:
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            model.appendRow(items)

        self.processed_table.setModel(model)
        self.processed_table.frozenTableView.setModel(model)
        self.processed_table.model().layoutChanged.emit()
        self.processed_table.frozenTableView.model().layoutChanged.emit()
        logger.debug(f"Show processed pivot data took {time.time() - start_time:.3f} seconds")

    def open_search_window(self):
        """Open a dialog for search input with improved UI"""
        start_time = time.time()
        if self.app.get_data() is None:
            QMessageBox.warning(self, "Warning", "No data to search!")
            return

        if self.search_window and self.search_window.isVisible():
            self.search_window.close()

        self.search_window = QDialog(self)
        self.search_window.setWindowTitle("Search Pivot Table")
        self.search_window.setFixedSize(350, 200)
        self.search_window.setStyleSheet(global_style)
        layout = QVBoxLayout(self.search_window)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        layout.addWidget(QLabel("Search in Pivot Table:", font=QFont("Inter", 12, QFont.Weight.Bold)))
        search_entry = QLineEdit()
        search_entry.setPlaceholderText("Enter search term...")
        search_entry.setToolTip("Enter text to search in the pivot table")
        search_entry.textChanged.connect(lambda text: setattr(self, 'search_var', text))
        search_entry.textChanged.connect(lambda: QTimer.singleShot(0, self.show_processed_data))
        layout.addWidget(search_entry)

        button_layout = QHBoxLayout()
        search_btn = QPushButton("🔍 Search")
        search_btn.setToolTip("Perform the search")
        search_btn.clicked.connect(self.show_processed_data)
        button_layout.addWidget(search_btn)

        close_btn = QPushButton("Close")
        close_btn.setToolTip("Close the search window")
        close_btn.clicked.connect(self.search_window.close)
        button_layout.addWidget(close_btn)

        layout.addLayout(button_layout)
        layout.addStretch()

        self.search_window.show()
        search_entry.setFocus()
        logger.debug(f"Open search window took {time.time() - start_time:.3f} seconds")

    def open_filter_window(self):
        """Open a dialog to select filter values with improved UI"""
        start_time = time.time()
        df = self.app.get_data()
        if df is None:
            QMessageBox.warning(self, "Warning", "No data to filter!")
            return

        dialog = FilterDialog(
            self, self.filter_values, self.filter_field,
            self.solution_label_order, self.element_order,
            self.show_processed_data
        )
        dialog.accepted.connect(self.show_processed_data)
        dialog.exec()
        self.filter_field = dialog.filter_combo.currentText()
        logger.debug(f"Open filter window took {time.time() - start_time:.3f} seconds")

    def save_processed_excel(self):
        """Save processed pivot table to Excel with styles"""
        start_time = time.time()
        df = self.get_filtered_data()
        if df is None:
            QMessageBox.warning(self, "Warning", "No data to save!")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", "", "Excel Files (*.xlsx)"
        )
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Processed Pivot Table"

                header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                first_column_fill = PatternFill(start_color="FFF5E4", end_color="FFF5E4", fill_type="solid")
                odd_row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                header_font = Font(name="Inter", size=12, bold=True)
                cell_font = Font(name="Inter", size=12)
                cell_alignment = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin")
                )

                headers = list(df.columns)
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15

                decimal_places = int(self.decimal_combo.currentText())
                for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                    fill = even_row_fill if (row_idx - 1) % 2 == 0 else odd_row_fill
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        try:
                            cell.value = round(float(value), decimal_places)
                            cell.number_format = f"0.{'0' * decimal_places}"
                        except (ValueError, TypeError):
                            cell.value = value
                        cell.font = cell_font
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                        if col_idx == 1:
                            cell.fill = first_column_fill
                        else:
                            cell.fill = fill

                wb.save(file_path)
                QMessageBox.information(self, "Success", "Processed pivot table saved successfully!")
                logger.debug(f"Save processed excel took {time.time() - start_time:.3f} seconds")

                if QMessageBox.question(self, "Open File", "Would you like to open the saved Excel file?") == QMessageBox.StandardButton.Yes:
                    try:
                        system = platform.system()
                        if system == "Windows":
                            os.startfile(file_path)
                        elif system == "Darwin":
                            os.system(f"open {file_path}")
                        else:
                            os.system(f"xdg-open {file_path}")
                        logger.debug(f"Opened file: {file_path}")
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to open file: {str(e)}")
                        logger.error(f"Failed to open file: {str(e)}")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save: {str(e)}")
                logger.error(f"Failed to save: {str(e)}")

    def reset_cache(self):
        """Reset cached data, orders, filters, and search"""
        self.last_filtered_data = None
        self._last_cache_key = None
        self.solution_label_order = None
        self.element_order = None
        self.column_widths = {}
        self.filter_values = {}
        self.search_var = ""
        self.data_hash = None
        logger.debug("ResultsFrame cache, orders, filters, and search reset")